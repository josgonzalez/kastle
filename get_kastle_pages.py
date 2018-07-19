from wordpress_xmlrpc import Client, WordPressPost, WordPressPage, WordPressTerm
from wordpress_xmlrpc.methods.posts import GetPosts, NewPost
from wordpress_xmlrpc.methods.users import GetUserInfo
from wordpress_xmlrpc.methods import posts, taxonomies  
from openpyxl import Workbook

client = Client('https://kastle.kippla.org/xmlrpc.php', 'Jose Test', 'lakai219')
increment = 100
offset = 0
#create spreadsheet workbook
wb = Workbook()
#create sheet at first position
ws = wb.create_sheet("Kastle Dashboard Pages", 0)
ws.cell(row=1, column=1).value = 'Title'
ws.cell(row=1, column=2).value = 'Category'
ws.cell(row=1, column=3).value = 'Owner'
ws.cell(row=1, column=4).value = 'Tags'
ws.cell(row=1, column=5).value = 'User Access'
ws.cell(row=1, column=6).value = 'Subject Type'
ws.cell(row=1, column=7).value = 'Target Audience'
tags = []
target_audiences = []
user_access = []
subject_types = []
i = 2

pages = client.call(posts.GetPosts({'number': increment, 'offset': offset, 'post_type': 'page'}, results_class=WordPressPage))

for page in pages :
    terms = page.terms
    for term in terms:
        if term.taxonomy == 'category':
            cat = term.name
            if cat != 'Non Dashboard':
                ws.cell(row=i, column=1).value = page.title
                ws.cell(row=i, column=2).value = cat
                for term in terms:
                    if term.taxonomy == 'owner':
                        ws.cell(row=i, column=3).value = term.name
                    elif term.taxonomy == 'post_tag':
                        tags.append(term.name)
                    elif term.taxonomy == 'target_audience':
                        target_audiences.append(term.name)
                    elif term.taxonomy == 'user_access':
                        user_access.append(term.name)
                    elif term.taxonomy == 'subject_type':
                        subject_types.append(term.name)
                ws.cell(row=i, column=4).value = ' '.join(tags)
                ws.cell(row=i, column=5).value = ' '.join(user_access)
                ws.cell(row=i, column=6).value = ' '.join(subject_types)
                ws.cell(row=i, column=7).value = ' '.join(target_audiences)
                tags = []
                target_audiences = []
                user_access = []
                subject_types = []
                i = i + 1

wb.save('Kastle Dashboards.xlsx')
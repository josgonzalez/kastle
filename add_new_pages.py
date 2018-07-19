from wordpress_xmlrpc import Client, WordPressPost, WordPressPage, WordPressTerm
from wordpress_xmlrpc.methods.posts import GetPosts, NewPost
from wordpress_xmlrpc.methods.users import GetUserInfo
from wordpress_xmlrpc.methods import posts, taxonomies, media  
from wordpress_xmlrpc.compat import xmlrpc_client
from openpyxl import load_workbook

#IDs for all parent pages
parentPages = {
    'Planning & Accountability': 752, 
    'CA Dashboard': 4311,    

    'Academics': 4157,
    'Grades': 1977,
    'Illuminate Assessments': 758, 
    'MAP - All Seasons': 736,
    'Fall MAP': 1820,
    'Spring MAP': 785,
    'Winter MAP': 733,
    'Reading': 761,
    'SBAC - Public': 1857,
    'SBAC - KLA': 1842,

    'Attendance & Behavior': 3923,
    'Behavior': 3437,
    'Attendance': 3926,

    'Enrollment': 1894,
    'Attrition': 1932,
    'Demographics': 721,
    'Recruitment & Enrollment': 718,

    'Student Groups & Programs': 3433,
    'EL': 743,
    'SPED': 4076,

    'Surveys': 1916, 
    'Family': 773,
    'TNTP': 767,
    'Student': 770,

    'KTC': 782,

    'Big KIPPsters': 4263,
    'HR': 3440,
}  

#authorization for wordpress
client = Client('https://kastle.kippla.org/xmlrpc.php', 'Jose Test', 'lakai219')
#create wordpress page to be added
page = WordPressPage()
#load spreadsheet with dashboard information
wb = load_workbook('Tableau Dashboards.xlsx')
#load sheet in spreadsheet
ws = wb['Tableau Dashboards']

#function to determine access level
def getAccessLevel(x) :
    accessLevels = None
    if (x == "All Members"):
        accessLevels = ["all users"] 
    if (x == "SST & School Leaders" or x == "SST and School Leaders Only") :
        accessLevels = ["school leaders"] 
    if (x == "SST Only" or x == "SST") :
        accessLevels = ["sst"] 
    return accessLevels

#function to determine with extra icons to insert
def getExtraIcons(haveFilters, haveActions) :
    extraIcons = None
    if (haveFilters == "Y" and haveActions == "Y") :
        extraIcons = ["clickable", "school filters"]
    elif (haveFilters == "Y" and haveActions == "N") :
        extraIcons = ["school filters"]
    elif (haveFilters == "N" and haveActions == "Y") :
        extraIcons = ["clickable"]
    else : 
        extraIcons = []
    return extraIcons

cat = None

#iterate throught the sheet
for i in range (123, 136):
    #get screenshot saved on computer and upload to wordpress
    filename = '/Users/jgonzalez/Desktop/view_preview_images/' + ws.cell(row=i, column=1).value + '_' + ws.cell(row=i, column=4).value + '.jpg'
    data = {
        'name':  ws.cell(row=i, column=1).value + '_' + ws.cell(row=i, column=4).value + '.jpg',
        'type': 'image/jpg', 
    }
    with open(filename, 'rb') as img:
        data['bits'] = xmlrpc_client.Binary(img.read())
    response = client.call(media.UploadFile(data))
    thumbnail_id = response['id']


    page.title = ws.cell(row=i, column=4).value
    page.parent_id = parentPages[ws.cell(row=i, column=2).value]
    cat = ws.cell(row=i, column=2).value

    #if there is a subcategory then make that the parent
    if ws.cell(row=i, column=3).value:
        page.parent_id = parentPages[ws.cell(row=i, column=3).value]
        cat = ws.cell(row=i, column=3).value
        
    #various taxonomies for page
    page.excerpt = ws.cell(row=i, column=5).value
    page.terms_names = {
        'category': [cat],
        'subject_type': ws.cell(row=i, column=7).value.splitlines(),
        'target_audience': ws.cell(row=i, column=8).value.splitlines(),
        'user_access': getAccessLevel(ws.cell(row=i, column=9).value),
        'owner': [ws.cell(row=i, column=6).value],
        'update_freq': [ws.cell(row=i, column=10).value],
        'extra_icons' : getExtraIcons(ws.cell(row=i, column=11).value, ws.cell(row=i, column=12).value)
    }

    #custom fields for page
    page.custom_fields = []
    page.custom_fields.append({
        'key': 'dashboard_link', 'value': ws.cell(row=i, column=16).value
    })
    page.custom_fields.append({
        'key': 'dashboard_tips', 'value': ws.cell(row=i, column=14).value
    })
    page.custom_fields.append({
        'key': 'context', 'value': ws.cell(row=i, column=15).value
    })

    #thumbnail id of the image we just uploaded
    page.thumbnail = thumbnail_id
    #publish page
    page.post_status = 'publish'
    #create page
    page.id = client.call(posts.NewPost(page))

    print ("Added " + cat + ": " + ws.cell(row=i, column=4).value)
    print ("")


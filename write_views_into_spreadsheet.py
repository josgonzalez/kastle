import tableauserverclient as TSC
from openpyxl import Workbook

#authentication to tableau server
tableau_auth = TSC.TableauAuth('kippndc\josgonzalez', 'Lakai219', 'KIPPLA')
#server
server = TSC.Server('https://tableau.kipp.org/', use_server_version=True)
#create spreadsheet workbook
wb = Workbook()
#create sheet at first position
ws = wb.create_sheet("Tableau Dashboards", 0)

ws.cell(row=1, column=1).value = 'Workbook'
ws.cell(row=1, column=2).value = 'Category'
ws.cell(row=1, column=3).value = 'Subcategory'
ws.cell(row=1, column=4).value = 'Dashboard Name'
ws.cell(row=1, column=5).value = 'Description'
ws.cell(row=1, column=6).value = 'Owner'
ws.cell(row=1, column=7).value = 'Data Subject Type'
ws.cell(row=1, column=8).value = 'Target Audience'
ws.cell(row=1, column=9).value = 'Team Member'
ws.cell(row=1, column=10).value = 'Data Update Frequency'
ws.cell(row=1, column=11).value = 'School Filter?'
ws.cell(row=1, column=12).value = 'Actions?'
ws.cell(row=1, column=13).value = 'Tags'
ws.cell(row=1, column=14).value = 'Dashboard Tips'
ws.cell(row=1, column=15).value = 'Context and Definitions'
ws.cell(row=1, column=16).value = 'Embed Code'

with server.auth.sign_in(tableau_auth): 
    req_options = TSC.RequestOptions()
    req_options.page_size(1000)
    req_options.sort.add(TSC.Sort(TSC.RequestOptions.Field.Name, TSC.RequestOptions.Direction.Asc))
    all_views, pagination_item = server.views.get(req_options)

    i = 2
    for view in all_views :
        owner = server.users.get_by_id(view.owner_id)
        workbook = server.workbooks.get_by_id(view.workbook_id)
        #only get views in KASTLE project
        if (workbook.project_name == 'KASTLE') :
            ws.cell(row=i, column=1).value = workbook.name
            #split workbook name into array of strings
            categories = workbook.name.split('_')
            #if there are 3 string then there is a sub category
            if len(categories) == 3 :
                ws.cell(row=i, column=2).value = categories[0]
                ws.cell(row=i, column=3).value = categories[1]
            #if there are 2 strings then there is no sub category
            if len(categories) == 2:
                ws.cell(row=i, column=2).value = categories[0]
            ws.cell(row=i, column=4).value = view.name
            ws.cell(row=i, column=6).value = owner.fullname
            #embedd code determined by workbook and view name
            ws.cell(row=i, column=16).value = "<script type='text/javascript' src='https://tableau.kipp.org/javascripts/api/viz_v1.js'></script><div class='tableauPlaceholder' style='width: 1200px; height: 100px;'><object class='tableauViz' width='1200' height='1000' style='display:none;'><param name='host_url' value='https%3A%2F%2Ftableau.kipp.org%2F' /> <param name='embed_code_version' value='3' /> <param name='site_root' value='&#47;t&#47;KIPPLA' /><param name='name' value='" + workbook.name.replace(' ', '').replace('&', '').replace('/','').replace('(', '').replace(')','') + "&#47;" + view.name.replace(' ', '').replace('&', '').replace('/','').replace('(', '').replace(')','') + "' /><param name='tabs' value='no' /><param name='toolbar' value='yes' /><param name='showAppBanner' value='false' /></object></div>"
            print (i - 1)   
            i = i + 1
            print ("Inserted " + workbook.name + " " + view.name)

#save spreadsheet
wb.save('Tableau Dashboards.xlsx')

